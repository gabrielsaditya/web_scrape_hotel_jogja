from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import pyperclip
import re

# Fungsi tunggu elemen dengan visibility
def tunggu_elemen(by, selector, timeout=15):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, selector))
    )

# Fungsi tunggu elemen clickable
def tunggu_elemen_clickable(by, selector, timeout=15):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, selector))
    )

# Fungsi ambil rating
def get_rating(driver):
    try:
        spans = driver.find_elements(By.XPATH, 
            "//span[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'star hotel') or contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'hotel bintang')]"
        )
        for span in spans:
            text = span.text.strip()
            match = re.search(r"(\d+)[-\s]*star hotel", text, re.IGNORECASE)
            if match:
                return match.group(1)
            match = re.search(r"hotel bintang\s+(\d+)", text, re.IGNORECASE)
            if match:
                return match.group(1)
    except Exception as e:
        print("❌ Error in get_rating:", e)
    return ""

# Fungsi parsing alamat
def parse_alamat(alamat_lengkap):
    parts = [x.strip() for x in alamat_lengkap.split(',')]
    jalan = parts[0] if len(parts) > 0 else ''
    kelurahan = parts[1] if len(parts) > 1 else ''
    kecamatan = parts[2] if len(parts) > 2 else ''
    kabupaten = ''
    for part in parts:
        if 'Kabupaten' in part or 'Kota' in part:
            kabupaten = part
            break
    return jalan, kelurahan, kecamatan, kabupaten

# Setup driver
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
try:
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(3)
except Exception as e:
    print(f"🚫 Gagal inisialisasi driver: {e}")
    exit()

# Load workbook
file_excel = 'Copy of HOSPIT BIZ YOGYAKARTA (1).xlsx'
wb = openpyxl.load_workbook(file_excel)
ws = wb['HOSPIT BIZ YOGYAKARTA']

# Tambah header jika belum ada
headers = ["", "Nama Hotel", "", "", "", "", "", "Alamat Lengkap", "Telepon", "Tikor", "ShareLok", "Rating Bintang", "Jalan", "Kelurahan", "Kecamatan", "Kabupaten"]
if ws.cell(row=1, column=8).value != "Alamat Lengkap":
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx).value = header

# Buka Google Maps dan tunggu search box muncul
driver.get("https://www.google.com/maps")
tunggu_elemen(By.ID, "searchboxinput")

# Loop data hotel, hanya proses baris jika kolom "Alamat Lengkap" (kolom 8) kosong
for row in range(2, ws.max_row + 1):
    # Skip baris jika "Alamat Lengkap" sudah terisi
    if ws.cell(row=row, column=8).value not in (None, "", " "):
        print(f"Baris {row} sudah terisi, skip...")
        continue

    nama_hotel = ws.cell(row=row, column=2).value
    if not nama_hotel:
        continue

    print(f"\n🔍 Mencari: {nama_hotel}")

    # Coba escape overlay/popup jika ada
    try:
        body = tunggu_elemen(By.TAG_NAME, "body")
        body.send_keys(Keys.ESCAPE)
        time.sleep(1)
    except Exception:
        pass

    # Isi search box dengan aman
    try:
        search_box = tunggu_elemen_clickable(By.ID, "searchboxinput")
        # Bersihkan nilai search box menggunakan JavaScript
        driver.execute_script("arguments[0].value = '';", search_box)
        time.sleep(0.5)
        search_box.send_keys(nama_hotel)
        search_box.send_keys(Keys.ENTER)
    except Exception as e:
        print("❌ Error saat input pencarian:", e)
        continue

    # Tunggu hasil pencarian muncul
    time.sleep(6)
    driver.execute_script("window.scrollBy(0, 500);")
    time.sleep(1)

    # Ambil alamat
    try:
        alamat_elem = tunggu_elemen(By.XPATH, '//button[contains(@aria-label, "Alamat") or contains(@aria-label, "Address")]')
        alamat = alamat_elem.get_attribute("aria-label").replace("Alamat: ", "").replace("Address: ", "")
        if not alamat:
            alamat = alamat_elem.text
    except Exception:
        alamat = "-"
    print(f"📍 {alamat}")

    # Ambil telepon
    try:
        telepon_elem = tunggu_elemen(By.XPATH, '//button[contains(@aria-label, "Telepon") or contains(@aria-label, "Phone")]')
        telepon = telepon_elem.get_attribute("aria-label").replace("Telepon: ", "").replace("Phone: ", "")
        if not telepon:
            telepon = telepon_elem.text
    except Exception:
        telepon = "-"

    # Ambil rating
    rating_angka = get_rating(driver)
    print(f"⭐ Rating: {rating_angka}")

    # Ambil link share lokasi
    try:
        share_btn = tunggu_elemen_clickable(By.XPATH, "//button[contains(., 'Bagikan')]")
        share_btn.click()
        time.sleep(2)

        copy_link_btn = tunggu_elemen_clickable(By.XPATH, "//button[contains(., 'Salin link') or contains(., 'Copy link')]")
        copy_link_btn.click()
        time.sleep(2)  # Beri waktu agar clipboard terisi

        share_link = pyperclip.paste()
        print(f"✅ Share link disalin: {share_link}")

        # Tutup overlay share jika ada
        try:
            close_btn = tunggu_elemen_clickable(By.XPATH, '//button[contains(@aria-label, "Tutup")]')
            close_btn.click()
            time.sleep(1)
        except Exception:
            pass
    except Exception as e:
        share_link = "-"
        print("❌ Gagal salin share link:", e)

    # Simpan data ke Excel
    ws.cell(row=row, column=8).value = alamat
    ws.cell(row=row, column=9).value = telepon
    ws.cell(row=row, column=11).value = share_link
    ws.cell(row=row, column=3).value = rating_angka

    # Pisahkan alamat menjadi beberapa bagian
    jalan, kelurahan, kecamatan, kabupaten = parse_alamat(alamat)
    ws.cell(row=row, column=4).value = jalan
    ws.cell(row=row, column=5).value = kelurahan
    ws.cell(row=row, column=6).value = kecamatan
    ws.cell(row=row, column=7).value = kabupaten
    print(f"🏷️ Alamat dipisah: {jalan}, {kelurahan}, {kecamatan}, {kabupaten}")

    wb.save(file_excel)
    print(f"💾 Data disimpan untuk: {nama_hotel}")

driver.quit()
